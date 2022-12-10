from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
import csv
import math
from datetime import datetime
from prettytable import PrettyTable
from typing import *

NAME = 0
SALARY_FROM = 1
SALARY_TO = 2
SALARY_CURRENCY = 3
AREA_NAME = 4
PUBLISHED_AT = 5

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


class Report:
    def __init__(self, filename, name):
        self.filename = filename
        self.name = name
        self.years = list(range(2007, 2023))
        self.years_sums = {}
        self.years_length = {}
        self.years_sums_cur = {}
        self.years_length_cur = {}
        self.cities = []
        self.cities_sums = {}
        self.cities_length = {}
        self.vacancies_length = 0
        self.ans_cities_sums = {}
        self.cities_partitions = {}
        self.read_file()
        self.calculate_file()
        self.Wb = Workbook()

    def read_file(self):
        first = False
        with open(self.filename, encoding="utf-8") as file:
            reader = csv.reader(file)
            for row in reader:
                if not first:
                    first = True
                    NAME = row.index("name")
                    SALARY_FROM = row.index("salary_from")
                    SALARY_TO = row.index("salary_to")
                    SALARY_CURRENCY = row.index("salary_currency")
                    AREA_NAME = row.index("area_name")
                    PUBLISHED_AT = row.index("published_at")
                else:
                    my_row = row.copy()
                    if all(my_row):
                        cur_year = int(row[PUBLISHED_AT].split("-")[0])
                        cur_salary = (int(float(row[SALARY_TO])) + int(float(row[SALARY_FROM]))) * currency_to_rub[
                            row[SALARY_CURRENCY]] // 2
                        cur_name = row[NAME]
                        cur_city = row[AREA_NAME]
                        self.years_sums[cur_year] = self.years_sums.get(cur_year, 0) + cur_salary
                        self.years_length[cur_year] = self.years_length.get(cur_year, 0) + 1
                        if name in cur_name:
                            self.years_sums_cur[cur_year] = self.years_sums_cur.get(cur_year, 0) + cur_salary
                            self.years_length_cur[cur_year] = self.years_length_cur.get(cur_year, 0) + 1
                        if cur_city not in self.cities:
                            self.cities.append(cur_city)
                        self.cities_sums[cur_city] = self.cities_sums.get(cur_city, 0) + cur_salary
                        self.cities_length[cur_city] = self.cities_length.get(cur_city, 0) + 1
                        self.vacancies_length += 1

    def calculate_file(self):
        for i in self.years:
            if self.years_sums.get(i, None):
                self.years_sums[i] = int(self.years_sums[i] // self.years_length[i])
            if self.years_sums_cur.get(i, None):
                self.years_sums_cur[i] = int(self.years_sums_cur[i] // self.years_length_cur[i])

        for i in self.cities:
            self.cities_sums[i] = int(self.cities_sums[i] // self.cities_length[i])
        interesting_cities = [city for city in self.cities if self.cities_length[city] >= self.vacancies_length // 100]
        self.ans_cities_sums = {key: self.cities_sums[key] for key in
                                sorted(interesting_cities, key=lambda x: self.cities_sums[x], reverse=True)[:10]}
        self.cities_partitions = {key: float("{:.4f}".format(self.cities_length[key] / self.vacancies_length)) for key
                                  in
                                  sorted(interesting_cities,
                                         key=lambda x: self.cities_length[x] / self.vacancies_length,
                                         reverse=True)[:10]}

    def print_file(self):
        print("Динамика уровня зарплат по годам:", self.years_sums)
        print("Динамика количества вакансий по годам:", self.years_length)
        if not len(self.years_sums_cur):
            self.years_sums_cur[2022] = 0
        print("Динамика уровня зарплат по годам для выбранной профессии:", self.years_sums_cur)
        if not len(self.years_length_cur):
            self.years_length_cur[2022] = 0
        print("Динамика количества вакансий по годам для выбранной профессии:", self.years_length_cur)
        print("Уровень зарплат по городам (в порядке убывания):", self.ans_cities_sums)
        print("Доля вакансий по городам (в порядке убывания):", self.cities_partitions)

    def generate_excel(self):
        self.years_stat_sheet = self.Wb.create_sheet(title="Статистика по годам")
        self.cities_stat_sheet = self.Wb.create_sheet(title="Статистика по городам")
        self.Wb.remove(self.Wb["Sheet"])
        sd = Side(border_style='thin', color="000000")
        self.border = Border(right=sd, top=sd, bottom=sd, left=sd)
        self.header_alignment = Alignment(horizontal='left')
        self.data_alignment = Alignment(horizontal='right')
        self.cities_stat_sheet["a1"] = 12
        self.report_years()
        self.report_cities()
        self.fit_cells()
        self.Wb.save('report.xlsx')

    def report_years(self):
        headers = ["Год", "Средняя зарплата", "Средняя зарплата - " + self.name,
                   "Количество вакансий", "Количество вакансий - " + self.name]
        self.set_headers(self.years_stat_sheet, headers)

        matrix = []
        for row in range(len(self.years_sums)):
            key = list(self.years_sums.keys())[row]
            appendable = [key, self.years_sums[key], self.years_sums_cur[key], self.years_length[key],
                          self.years_length_cur[key]]
            matrix.append(appendable)

        self.fill_matrix(self.years_stat_sheet, matrix, offset=(0, 1))

    def fill_matrix(self, sheet, matrix, offset=(0, 0)):
        for row in range(len(matrix)):
            for col in range(len(matrix[0])):
                address = f"{get_column_letter(col + 1 + offset[0])}{row + 1 + offset[1]}"
                sheet[address] = matrix[row][col]
                sheet[address].border = self.border
                sheet[address].alignment = self.data_alignment
                sheet.column_dimensions[get_column_letter(col + 1)].auto_size = 1

    def set_headers(self, sheet, headers, offset=(0, 0)):
        for col in range(0, len(headers)):
            address = f"{get_column_letter(col + 1 + offset[0])}{1 + offset[1]}"
            sheet[address] = headers[col]
            sheet[address].border = self.border
            sheet[address].alignment = self.header_alignment
            sheet[address].font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(col + 1)].auto_size = 1

    def fit_cells(self):
        for sheet_name in self.Wb.sheetnames:
            sheet = self.Wb[sheet_name]
            for col in range(1, sheet.max_column + 1):
                width = None
                for row in range(1, sheet.max_row + 1):
                    value = sheet[f"{get_column_letter(col)}{row}"].value
                    if value is not None and (width is None or len(str(value)) > width):
                        width = len(str(value))
                if width is not None:
                    sheet.column_dimensions[f"{get_column_letter(col)}"].width = width + 2
                else:
                    sheet.column_dimensions[f"{get_column_letter(col)}"].width = + 2

    def report_cities(self):
        headers_payment = ["Город", "Уровень зарплат"]
        headers_percent = ["Город", "Доля вакансий"]
        self.set_headers(self.cities_stat_sheet, headers_payment)
        self.set_headers(self.cities_stat_sheet, headers_percent, (3, 0))

        self.data_alignment = Alignment(horizontal='left')
        self.fill_matrix(self.cities_stat_sheet, [[i] for i in self.ans_cities_sums.keys()], offset=(0, 1))
        matrix = {key: f"{(val * 10000) // 1 / 100}%" for key, val in self.cities_partitions.items()}
        self.fill_matrix(self.cities_stat_sheet, [[i] for i in list(matrix.keys())], offset=(3, 1))
        self.data_alignment = Alignment(horizontal='right')
        self.fill_matrix(self.cities_stat_sheet, [[i] for i in list(self.ans_cities_sums.values())], offset=(1, 1))
        self.fill_matrix(self.cities_stat_sheet, [[i] for i in list(matrix.values())], offset=(4, 1))

    def generate_image(self):
        matplotlib.rc("font", size=8)
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)
        width = 0.3
        x = np.arange(len(self.years_sums.keys()))
        payment1 = ax1.bar(x - width / 2, self.years_sums.values(), width, label="средняя з/п")
        payment2 = ax1.bar(x + width / 2, self.years_sums_cur.values(), width, label=f"з/п {self.name}")

        ax1.grid(True, axis="y")
        ax1.set_title("Уровень зарплат по годам")
        ax1.set_xticks(np.arange(len(self.years_sums.keys())), self.years_sums.keys(), rotation=90)
        ax1.bar_label(payment1, fmt="")
        ax1.bar_label(payment2, fmt="")
        ax1.legend(prop={"size": 6})

        ax2.grid(True, axis="y")
        ax2.set_title("Количество вакансий по годам")
        x = np.arange(len(self.years_sums.keys()))
        ax2.set_xticks(x, self.years_sums.keys(), rotation=90)
        vac1 = ax2.bar(x - width / 2, self.years_sums.values(), width, label="Количество вакансий")
        vac2 = ax2.bar(x + width / 2, self.years_sums_cur.values(), width, label=f"Количество вакансий\n{self.name}")
        ax2.bar_label(vac1, fmt="")
        ax2.bar_label(vac2, fmt="")
        ax2.legend(prop={"size": 6})

        ax3.grid(True, axis="x")
        y = np.arange(len(list(self.ans_cities_sums.keys())))
        ax3.set_yticks(y, map(lambda s: s.replace(" ", "\n").replace("-", "\n"), self.ans_cities_sums.keys()))
        ax3.invert_yaxis()
        ax3.barh(y, self.ans_cities_sums.values())
        ax3.set_title("Уровень зарплат по городам")

        ax4.set_title("Доля вакансий по городам")
        other = 1 - sum(self.cities_partitions.values())
        ax4.pie([other] + list(self.cities_partitions.values()),
                labels=["Другие"] + list(self.cities_partitions.keys()), startangle=0)

        fig.tight_layout(pad=0.4, w_pad=0.5, h_pad=1.0)
        plt.savefig("graph.png")


def prettify_val(val):
    if type(val) == list:
        val = "\n".join(val)
    val = str(val)
    if len(val) > 100:
        val = val[:100] + "..."
    return val


def parse_money(amount):
    nseq = []
    seq = list(reversed(list(str(amount))))
    for i in range(len(seq)):
        nseq.append(seq[i])
        if i % 3 == 2:
            nseq.append(" ")
    return "".join(reversed(nseq)).strip()


def try_parse(val):
    if val == math.nan:
        return "nan"
    if val == "True":
        return "TRUE"
    if val == "False":
        return "FALSE"
    try:
        val = int(float(val))
    finally:
        return str(val)


def skills_filter(vac, *args):
    for skill in args[1].split(", "):
        if skill not in vac["key_skills"]:
            return False
    return True


def salary_filter(vac, *args):
    return int(vac["salary_from"]) <= int(args[1]) <= int(vac["salary_to"])


def publish_filter(vac, *args):
    return datetime.strptime(".".join(vac["published_at"].split("T")[0].split("-")[::-1]),
                             "%d.%m.%Y") == datetime.strptime(args[1], "%d.%m.%Y")


def parameter_filter(vac, *args):
    return DIC_PARAM[vac[dic_terms[args[0]]]] == args[1]


def premium_filter(vac, *args):
    return dic_joke[vac["premium"]] == args[1]


def simple_parameter_filter(vac, *args):
    return vac[dic_terms[args[0]]] == args[1]


def get_filter(func, *args):
    def parameter_func(vac):
        return func(vac, *args)

    return parameter_func


DIC_FILTER = {"Навыки": skills_filter,
              "Оклад": salary_filter,
              "Дата публикации вакансии": publish_filter,
              "Опыт работы": parameter_filter,
              "Премиум-вакансия": premium_filter,
              "Идентификатор валюты оклада": parameter_filter,
              "Название": simple_parameter_filter,
              "Название региона": simple_parameter_filter,
              "Компания": simple_parameter_filter,
              "": lambda *x: True
              }
DIC_PARAM = {
    # exp
    "noExperience": "Нет опыта",
    "between1And3": "От 1 года до 3 лет",
    "between3And6": "От 3 до 6 лет",
    "moreThan6": "Более 6 лет",
    # money
    "AZN": "Манаты",
    "BYR": "Белорусские рубли",
    "EUR": "Евро",
    "GEL": "Грузинский лари",
    "KGS": "Киргизский сом",
    "KZT": "Тенге",
    "RUR": "Рубли",
    "UAH": "Гривны",
    "USD": "Доллары",
    "UZS": "Узбекский сум",
    # gross
    "Да": "Без вычета налогов",
    "Нет": "С вычетом налогов",
    "TRUE": "Без вычета налогов",
    "FALSE": "С вычетом налогов",
}
dic_joke = {
    "FALSE": "Нет",
    "False": "Нет",
    "Нет": "Нет",
    "TRUE": "Да",
    "True": "Да",
    "Да": "Да"
}
dic_trans = {"№": "№",
             "name": "Название",
             "description": "Описание",
             "key_skills": "Навыки",
             # "experience_id": "Опыт работы",
             "experience": "Опыт работы",
             "premium": "Премиум-вакансия",
             "employer_name": "Компания",
             # "salary_from": "Нижняя граница вилки оклада",
             # "salary_to": "Верхняя граница вилки оклада",
             # "salary_gross": "Оклад указан до вычета налогов",
             # "salary_currency": "Идентификатор валюты оклада",
             "salary": "Оклад",
             "area_name": "Название региона",
             # "published_at": "Дата и время публикации вакансии",
             "published_at_date": "Дата публикации вакансии"}
dic_terms = {
    "Название": "name",
    "Описание": "description",
    "Навыки": "key_skills",
    "Опыт работы": "experience_id",
    "Премиум-вакансия": "premium",
    "Компания": "employer_name",
    "Нижняя граница вилки оклада": "salary_from",
    "Верхняя граница вилки оклада": "salary_to",
    "Оклад указан до вычета налогов": "salary_gross",
    "Идентификатор валюты оклада": "salary_currency",
    "Оклад": "salary",
    "Название региона": "area_name",
    "Дата и время публикации вакансии": "published_at",
    "Дата публикации вакансии": "published_at_date"
}

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}
exp_list = ["Нет опыта",
            "От 1 года до 3 лет",
            "От 3 до 6 лет",
            "Более 6 лет"]
dic_sorters = {
    "": lambda v: True,
    "Название": lambda v: v.name,
    "Описание": lambda v: v.description,
    "Компания": lambda v: v.employer_name,
    "Название региона": lambda v: v.area_name,
    "Опыт работы": lambda v: exp_list.index(DIC_PARAM[v.experience_id]),
    "Премиум-вакансия": lambda v: dic_joke[v.premium],
    "Оклад": lambda v: (int(v.salary.salary_from) + int(v.salary.salary_to)) / 2 * currency_to_rub[
        v.salary.salary_currency],
    "Навыки": lambda v: len(v.key_skills) if type(v.key_skills) == list else 1,
    "Дата публикации вакансии": lambda v: [datetime.strptime(v.published_at, "%Y-%m-%dT%H:%M:%S%z")]
}



class DataSet:
    def __init__(self, file_name: str) -> None:
        self.file_name: str = file_name
        self.vacancies_objects: List[Vacancy] = []
        self.fill_vacancies()

    def read_file(self):
        keys = []
        values = []
        cnt = 1
        with open(self.file_name, encoding="utf-8") as file:
            reader = csv.reader(file)
            for row in reader:
                if not keys:
                    keys = ["№"] + row
                else:
                    my_row = row.copy()
                    if all(my_row):
                        values.append([str(cnt)] + [try_parse(i) for i in row])
                        cnt += 1
        values = list(filter(lambda x: "nan" not in x, values))
        if not len(keys):
            print("Пустой файл")
            exit(0)
        if not len(values):
            print("Нет данных")
            exit(0)
        return values, keys

    def fill_vacancies(self):
        reader, list_naming = self.read_file()
        for vacancy in reader:
            appendix = {}
            for i in range(len(vacancy)):
                append_item = vacancy[i].strip()
                tag_start = append_item.find("<")
                while tag_start != -1:
                    tag_end = append_item.find(">", tag_start)
                    append_item = append_item.replace(append_item[tag_start: tag_end + 1], "")
                    tag_start = append_item.find("<", tag_start)
                append_item = append_item.strip()
                while append_item.find("  ") != -1:
                    append_item = append_item.replace("  ", " ")
                if append_item.find("\n") != -1:
                    append_item = [" ".join(i.split()) for i in append_item.split("\n")]
                else:
                    append_item = " ".join(append_item.split())
                appendix[list_naming[i]] = append_item
            self.vacancies_objects.append(Vacancy(appendix))

    def filter_vacancies(self, filter_key, filter_val):
        filter_func = DIC_FILTER[filter_key]
        self.vacancies_objects = list(filter(lambda v: get_filter(filter_func, filter_key, *filter_val)(v.to_dict()),
                                             [vac for vac in self.vacancies_objects]))

    def sort_vacancies(self, name, reverse=False):
        self.vacancies_objects = sorted(self.vacancies_objects, key=dic_sorters[name], reverse=reverse)

    def prettify_vacancies(self, filter_key, filter_val, sort_name, reverse=False):
        self.filter_vacancies(filter_key, filter_val)
        self.sort_vacancies(sort_name, reverse)

    def print_vacancies(self, filter_key, filter_val, sort_name, dic_naming, reverse=False, row_indexes=None):
        if row_indexes is None:
            row_indexes = []
        self.prettify_vacancies(filter_key, filter_val, sort_name, reverse)
        pretty_vacancies = [vacancy.to_pretty_dict() for vacancy in self.vacancies_objects]
        if not row_indexes:
            row_indexes = [1, len(pretty_vacancies) + 1]
        if len(row_indexes) == 1:
            row_indexes = [row_indexes[0], len(pretty_vacancies) + 1]
        added, count = 0, 1
        output = PrettyTable(hrules=1, start=row_indexes[0] - 1, end=row_indexes[1] - 1)
        output.align = "l"
        for vac in pretty_vacancies:
            if not output.field_names:
                output.field_names = [name for name in dic_trans.keys() if vac.get(name, None) or name == "№"]
                dict_max = {}
                for key in output.field_names:
                    dict_max[dic_trans[key]] = 20
                output._max_width = dict_max
            vac["№"] = count
            addable = []
            for key in output.field_names:
                addable.append(prettify_val(vac.get(key)))
            count, added = count + 1, added + 1
            output.add_row(addable)
        output.field_names = [dic_trans[name] for name in output.field_names]
        if added != 0 and added >= row_indexes[0]:
            print(output.get_string(fields=list(dic_naming.values())))
        else:
            print("Ничего не найдено")


class Salary:
    def __init__(self, params):
        self.salary_from = params["salary_from"]
        self.salary_to = params["salary_to"]
        self.salary_gross = params["salary_gross"]
        self.salary_currency = params["salary_currency"]


class Vacancy:
    def __init__(self, params):
        self.name = params["name"]
        self.description = params["description"]
        self.key_skills = params["key_skills"]
        self.experience_id = params["experience_id"]
        self.premium = params["premium"]
        self.employer_name = params["employer_name"]
        self.salary = Salary(params)
        self.area_name = params["area_name"]
        self.published_at = params["published_at"]

    def to_dict(self):
        return {"name": self.name, "description": self.description, "key_skills": self.key_skills,
                "experience_id": self.experience_id, "premium": self.premium, "employer_name": self.employer_name,
                "salary_from": self.salary.salary_from, "salary_to": self.salary.salary_to,
                "salary_gross": self.salary.salary_gross, "salary_currency": self.salary.salary_currency,
                "area_name": self.area_name, "published_at": self.published_at}

    def to_pretty_dict(self):
        return {"name": self.name,
                "description": self.description,
                "key_skills": self.key_skills,
                "experience": DIC_PARAM[self.experience_id],
                "premium": dic_joke[self.premium],
                "employer_name": self.employer_name,
                "salary": f"{parse_money(self.salary.salary_from)} - " + f"{parse_money(self.salary.salary_to)} " +
                          f"({DIC_PARAM[self.salary.salary_currency]}) " + f"({DIC_PARAM[self.salary.salary_gross]})",
                "area_name": self.area_name,
                "published_at_date": ".".join(self.published_at.split("T")[0].split('-')[::-1])}


filename = input("Введите название файла: ")
name = input("Введите название профессии: ")

rep = Report(filename, name)
rep.print_file()
rep.generate_image()
