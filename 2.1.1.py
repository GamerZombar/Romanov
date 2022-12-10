from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import csv

NAME = 0
SALARY_FROM = 1
SALARY_TO = 2
SALARY_CURRENCY = 3
AREA_NAME = 4
PUBLISHED_AT = 5

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


class Report:
    def __init__(self, filename, name):
        self.filename = filename
        self.name = name
        self.years = list(range(2007, 2023))
        self.years_sums = {}
        self.years_length = {}
        self.years_sums_cur = {}
        self.years_length_cur = {}
        self.cities = []
        self.cities_sums = {}
        self.cities_length = {}
        self.vacancies_length = 0
        self.ans_cities_sums = {}
        self.cities_partitions = {}
        self.read_file()
        self.calculate_file()
        self.Wb = Workbook()

    def read_file(self):
        first = False
        with open(self.filename, encoding="utf-8") as file:
            reader = csv.reader(file)
            for row in reader:
                if not first:
                    first = True
                    NAME = row.index("name")
                    SALARY_FROM = row.index("salary_from")
                    SALARY_TO = row.index("salary_to")
                    SALARY_CURRENCY = row.index("salary_currency")
                    AREA_NAME = row.index("area_name")
                    PUBLISHED_AT = row.index("published_at")
                else:
                    my_row = row.copy()
                    if all(my_row):
                        cur_year = int(row[PUBLISHED_AT].split("-")[0])
                        cur_salary = (int(float(row[SALARY_TO])) + int(float(row[SALARY_FROM]))) * currency_to_rub[
                            row[SALARY_CURRENCY]] // 2
                        cur_name = row[NAME]
                        cur_city = row[AREA_NAME]
                        self.years_sums[cur_year] = self.years_sums.get(cur_year, 0) + cur_salary
                        self.years_length[cur_year] = self.years_length.get(cur_year, 0) + 1
                        if name in cur_name:
                            self.years_sums_cur[cur_year] = self.years_sums_cur.get(cur_year, 0) + cur_salary
                            self.years_length_cur[cur_year] = self.years_length_cur.get(cur_year, 0) + 1
                        if cur_city not in self.cities:
                            self.cities.append(cur_city)
                        self.cities_sums[cur_city] = self.cities_sums.get(cur_city, 0) + cur_salary
                        self.cities_length[cur_city] = self.cities_length.get(cur_city, 0) + 1
                        self.vacancies_length += 1

    def calculate_file(self):
        for i in self.years:
            if self.years_sums.get(i, None):
                self.years_sums[i] = int(self.years_sums[i] // self.years_length[i])
            if self.years_sums_cur.get(i, None):
                self.years_sums_cur[i] = int(self.years_sums_cur[i] // self.years_length_cur[i])

        for i in self.cities:
            self.cities_sums[i] = int(self.cities_sums[i] // self.cities_length[i])
        interesting_cities = [city for city in self.cities if self.cities_length[city] >= self.vacancies_length // 100]
        self.ans_cities_sums = {key: self.cities_sums[key] for key in
                                sorted(interesting_cities, key=lambda x: self.cities_sums[x], reverse=True)[:10]}
        self.cities_partitions = {key: float("{:.4f}".format(self.cities_length[key] / self.vacancies_length)) for key
                                  in
                                  sorted(interesting_cities,
                                         key=lambda x: self.cities_length[x] / self.vacancies_length,
                                         reverse=True)[:10]}

    def print_file(self):
        print("Динамика уровня зарплат по годам:", self.years_sums)
        print("Динамика количества вакансий по годам:", self.years_length)
        if not len(self.years_sums_cur):
            self.years_sums_cur[2022] = 0
        print("Динамика уровня зарплат по годам для выбранной профессии:", self.years_sums_cur)
        if not len(self.years_length_cur):
            self.years_length_cur[2022] = 0
        print("Динамика количества вакансий по годам для выбранной профессии:", self.years_length_cur)
        print("Уровень зарплат по городам (в порядке убывания):", self.ans_cities_sums)
        print("Доля вакансий по городам (в порядке убывания):", self.cities_partitions)

    def generate_excel(self):
        self.years_stat_sheet = self.Wb.create_sheet(title="Статистика по годам")
        self.cities_stat_sheet = self.Wb.create_sheet(title="Статистика по городам")
        self.Wb.remove(self.Wb["Sheet"])
        sd = Side(border_style='thin', color="000000")
        self.border = Border(right=sd, top=sd, bottom=sd, left=sd)
        self.header_alignment = Alignment(horizontal='left')
        self.data_alignment = Alignment(horizontal='right')
        self.cities_stat_sheet["a1"] = 12
        self.report_years()
        self.report_cities()
        self.fit_cells()
        self.Wb.save('report.xlsx')

    def report_years(self):
        headers = ["Год", "Средняя зарплата", "Средняя зарплата - " + self.name,
                   "Количество вакансий", "Количество вакансий - " + self.name]
        self.set_headers(self.years_stat_sheet, headers)

        matrix = []
        for row in range(len(self.years_sums)):
            key = list(self.years_sums.keys())[row]
            appendable = [key, self.years_sums[key], self.years_sums_cur[key], self.years_length[key],
                          self.years_length_cur[key]]
            matrix.append(appendable)

        self.fill_matrix(self.years_stat_sheet, matrix, offset=(0, 1))

    def fill_matrix(self, sheet, matrix, offset=(0, 0)):
        for row in range(len(matrix)):
            for col in range(len(matrix[0])):
                address = f"{get_column_letter(col + 1 + offset[0])}{row + 1 + offset[1]}"
                sheet[address] = matrix[row][col]
                sheet[address].border = self.border
                sheet[address].alignment = self.data_alignment
                sheet.column_dimensions[get_column_letter(col + 1)].auto_size = 1

    def set_headers(self, sheet, headers, offset=(0, 0)):
        for col in range(0, len(headers)):
            address = f"{get_column_letter(col + 1 + offset[0])}{1 + offset[1]}"
            sheet[address] = headers[col]
            sheet[address].border = self.border
            sheet[address].alignment = self.header_alignment
            sheet[address].font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(col + 1)].auto_size = 1

    def fit_cells(self):
        for sheet_name in self.Wb.sheetnames:
            sheet = self.Wb[sheet_name]
            for col in range(1, sheet.max_column + 1):
                width = None
                for row in range(1, sheet.max_row + 1):
                    value = sheet[f"{get_column_letter(col)}{row}"].value
                    if value is not None and (width is None or len(str(value)) > width):
                        width = len(str(value))
                if width is not None:
                    sheet.column_dimensions[f"{get_column_letter(col)}"].width = width + 2
                else:
                    sheet.column_dimensions[f"{get_column_letter(col)}"].width = + 2

    def report_cities(self):
        headers_payment = ["Город", "Уровень зарплат"]
        headers_percent = ["Город", "Доля вакансий"]
        self.set_headers(self.cities_stat_sheet, headers_payment)
        self.set_headers(self.cities_stat_sheet, headers_percent, (3, 0))

        self.data_alignment = Alignment(horizontal='left')
        self.fill_matrix(self.cities_stat_sheet, [[i] for i in self.ans_cities_sums.keys()], offset=(0, 1))
        matrix = {key: f"{(val * 10000) // 1 / 100}%" for key, val in self.cities_partitions.items()}
        self.fill_matrix(self.cities_stat_sheet, [[i] for i in list(matrix.keys())], offset=(3, 1))
        self.data_alignment = Alignment(horizontal='right')
        self.fill_matrix(self.cities_stat_sheet, [[i] for i in list(self.ans_cities_sums.values())], offset=(1, 1))
        self.fill_matrix(self.cities_stat_sheet, [[i] for i in list(matrix.values())], offset=(4, 1))


filename = input("Введите название файла: ")
name = input("Введите название профессии: ")

rep = Report(filename, name)
rep.print_file()
rep.generate_excel()
